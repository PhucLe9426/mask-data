import unittest
from io import BytesIO

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from app.backend import export_files


class ExportFileTests(unittest.TestCase):
    def setUp(self):
        self.title = "Tổng hợp hồ sơ"
        self.content = "Ứng viên có 5 năm kinh nghiệm.\nPhù hợp với vị trí kỹ sư."
        self.sources = [{"name": "ho-so.docx", "excerpt": "Kinh nghiệm phát triển phần mềm"}]

    def test_build_docx_contains_response_and_sources(self):
        data, media_type, filename = export_files.build_export(
            "docx", self.title, self.content, self.sources
        )
        document = Document(BytesIO(data))
        text = "\n".join(paragraph.text for paragraph in document.paragraphs)
        self.assertIn(self.title, text)
        self.assertIn("5 năm kinh nghiệm", text)
        self.assertIn("ho-so.docx", text)
        self.assertIn("wordprocessingml", media_type)
        self.assertTrue(filename.endswith(".docx"))

    def test_build_xlsx_has_summary_and_source_sheets(self):
        data, media_type, filename = export_files.build_export(
            "xlsx", self.title, self.content, self.sources
        )
        workbook = load_workbook(BytesIO(data), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Tổng hợp", "Nguồn"])
        self.assertEqual(workbook["Tổng hợp"]["A1"].value, self.title)
        self.assertIn("5 năm kinh nghiệm", workbook["Tổng hợp"]["A4"].value)
        self.assertEqual(workbook["Nguồn"]["A2"].value, "ho-so.docx")
        self.assertIn("spreadsheetml", media_type)
        self.assertTrue(filename.endswith(".xlsx"))

    def test_xlsx_treats_formula_like_output_as_text(self):
        data = export_files.build_xlsx("Báo cáo", "=HYPERLINK(\"bad\")", [])
        workbook = load_workbook(BytesIO(data), data_only=False)
        self.assertEqual(workbook["Tổng hợp"]["A4"].data_type, "s")
        self.assertTrue(workbook["Tổng hợp"]["A4"].value.startswith("'="))

    def test_build_pdf_is_readable(self):
        data, media_type, filename = export_files.build_export(
            "pdf", self.title, self.content, self.sources
        )
        reader = PdfReader(BytesIO(data))
        self.assertGreaterEqual(len(reader.pages), 1)
        self.assertTrue(data.startswith(b"%PDF"))
        self.assertEqual(media_type, "application/pdf")
        self.assertTrue(filename.endswith(".pdf"))

    def test_rejects_unknown_format(self):
        with self.assertRaises(ValueError):
            export_files.build_export("txt", self.title, self.content, [])


if __name__ == "__main__":
    unittest.main()
