import unittest
from io import BytesIO
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

from app.backend import file_reader


class FileReaderTests(unittest.TestCase):
    @staticmethod
    def xlsx_bytes(*, empty: bool = False) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ứng viên"
        if not empty:
            sheet.append(["Họ tên", "Kinh nghiệm"])
            sheet.append(["Lê Trọng Phúc", 5])
            formula_sheet = workbook.create_sheet("Đánh giá")
            formula_sheet["A1"] = "=1+1"
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def with_bogus_excel_dimension(content: bytes) -> bytes:
        source = BytesIO(content)
        output = BytesIO()
        with ZipFile(source) as incoming, ZipFile(output, "w", ZIP_DEFLATED) as outgoing:
            for item in incoming.infolist():
                data = incoming.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    data = data.replace(b'<dimension ref="A1:B2"/>', b'<dimension ref="A1:XFD1048576"/>')
                outgoing.writestr(item, data)
        return output.getvalue()

    def test_extract_utf8_text_and_sanitize_filename(self):
        name, text = file_reader.extract_text(
            "../../ghi-chu.txt",
            "Xin chào Việt Nam".encode(),
            max_bytes=1024,
            max_chars=100,
        )

        self.assertEqual(name, "ghi-chu.txt")
        self.assertEqual(text, "Xin chào Việt Nam")

    def test_reject_unsupported_extension(self):
        with self.assertRaisesRegex(file_reader.FileExtractionError, "chưa được hỗ trợ"):
            file_reader.extract_text(
                "payload.exe",
                b"not executable content",
                max_bytes=1024,
                max_chars=100,
            )

    def test_reject_empty_file(self):
        with self.assertRaisesRegex(file_reader.FileExtractionError, "đang trống"):
            file_reader.extract_text("empty.txt", b"", max_bytes=1024, max_chars=100)

    def test_reject_file_over_size_limit(self):
        with self.assertRaisesRegex(file_reader.FileExtractionError, "vượt quá giới hạn"):
            file_reader.extract_text(
                "large.txt",
                b"123456",
                max_bytes=5,
                max_chars=100,
            )

    def test_reject_extracted_text_over_character_limit(self):
        with self.assertRaisesRegex(file_reader.FileExtractionError, "sau khi trích xuất"):
            file_reader.extract_text(
                "long.txt",
                b"0123456789",
                max_bytes=100,
                max_chars=5,
            )

    def test_extract_xlsx_preserves_sheet_names_rows_and_formulas_as_text(self):
        content = self.xlsx_bytes()
        name, text = file_reader.extract_text(
            "../ho-so.xlsx",
            content,
            max_bytes=len(content) + 100,
            max_chars=10_000,
        )
        self.assertEqual(name, "ho-so.xlsx")
        self.assertIn("--- Sheet: Ứng viên ---", text)
        self.assertIn("Lê Trọng Phúc\t5", text)
        self.assertIn("--- Sheet: Đánh giá ---", text)
        self.assertIn("=1+1", text)

    def test_xlsx_ignores_incorrect_full_sheet_dimension(self):
        content = self.with_bogus_excel_dimension(self.xlsx_bytes())
        _, text = file_reader.extract_text(
            "dimension-ao.xlsx",
            content,
            max_bytes=len(content) + 100,
            max_chars=10_000,
        )
        self.assertIn("Lê Trọng Phúc\t5", text)

    def test_reject_empty_xlsx(self):
        content = self.xlsx_bytes(empty=True)
        with self.assertRaisesRegex(file_reader.FileExtractionError, "Không tìm thấy"):
            file_reader.extract_text(
                "empty.xlsx",
                content,
                max_bytes=len(content) + 100,
                max_chars=10_000,
            )

    def test_extract_legacy_xls(self):
        class FakeSheet:
            name = "Sheet cũ"
            nrows = 2

            @staticmethod
            def row_values(index):
                return [["Tên", "Điểm"], ["Phúc", 9]][index]

        class FakeWorkbook:
            def sheets(self):
                return [FakeSheet()]

            def release_resources(self):
                return None

        with patch.object(file_reader.xlrd, "open_workbook", return_value=FakeWorkbook()):
            name, text = file_reader.extract_text(
                "legacy.xls",
                b"legacy-binary",
                max_bytes=100,
                max_chars=1_000,
            )
        self.assertEqual(name, "legacy.xls")
        self.assertIn("Phúc\t9", text)


if __name__ == "__main__":
    unittest.main()
